#!/usr/bin/env python3
"""Script d'analyse des dossiers Outlook avec l'API Microsoft Graph"""

import os
import json
import time
from datetime import datetime
from typing import Dict, List, NamedTuple, Optional, TypedDict
import urllib.parse

import requests
import pandas as pd
from dotenv import load_dotenv
from azure.identity import ClientSecretCredential
from openpyxl.styles import Font, PatternFill, Alignment, Side, Border
from openpyxl.utils import get_column_letter
from tqdm import tqdm


class FolderAction(TypedDict):
    """Configuration des actions pour un dossier"""
    clean: bool
    archive: bool
    delete: bool
    min_size_mb: float
    max_age_days: int


class FolderConfig(TypedDict):
    """Configuration complète d'un dossier"""
    path: str
    clean: bool
    archive: bool
    delete: bool
    min_size_mb: float
    max_age_days: int
    last_action: Optional[str]
    last_action_date: Optional[str]


class FolderStats(NamedTuple):
    """Structure pour les statistiques d'un dossier"""
    id: str
    name: str
    path: str
    total_items: int
    total_size_mb: float
    total_attachments: int
    total_attachments_size_mb: float
    avg_size_mb: float
    avg_attachments_size_mb: float
    action_clean: bool
    action_archive: bool
    action_delete: bool
    action_min_size_mb: float
    action_max_age_days: int
    inherited_from: Optional[str]
    last_action: Optional[str]
    last_action_date: Optional[str]


class OutlookFolderAnalyzer:
    """Classe pour analyser les dossiers Outlook"""

    CONFIG_FILE = "folder_config.json"
    REPORT_FILE = "folder_analysis.xlsx"
    GRAPH_API_ENDPOINT = "https://graph.microsoft.com/v1.0"

    def __init__(self):
        """Initialisation de l'analyseur"""
        # Chargement des variables d'environnement
        load_dotenv()
        
        # Récupération des credentials
        self.client_id = os.getenv("OUTLOOK_CLIENT_ID")
        self.client_secret = os.getenv("OUTLOOK_CLIENT_SECRET")
        self.tenant_id = os.getenv("TENANT_ID")
        self.user_email = os.getenv("OUTLOOK_USER_EMAIL")

        if not all([
            self.client_id,
            self.client_secret,
            self.tenant_id,
            self.user_email
        ]):
            raise ValueError(
                "Veuillez configurer toutes les variables "
                "d'environnement requises"
            )

        # Authentification
        credentials = ClientSecretCredential(
            tenant_id=self.tenant_id,
            client_id=self.client_id,
            client_secret=self.client_secret
        )

        # Création de la session avec le token
        self.session = requests.Session()
        token = credentials.get_token("https://graph.microsoft.com/.default")
        self.session.headers.update({
            "Authorization": f"Bearer {token.token}",
            "Accept": "application/json"
        })

        # Chargement de la configuration
        self.config = self._load_config()

    def _load_config(self) -> Dict[str, FolderConfig]:
        """Charge la configuration des dossiers depuis le fichier JSON"""
        if os.path.exists(self.CONFIG_FILE):
            with open(self.CONFIG_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        return {}

    def _save_config(self) -> None:
        """Sauvegarde la configuration des dossiers"""
        with open(self.CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(self.config, f, indent=2, ensure_ascii=False)

    def get_folder_stats(self, folder: Dict, parent_path: str = "") -> FolderStats:
        """Récupère les statistiques d'un dossier"""
        folder_id = folder["id"]
        folder_name = folder["displayName"]
        
        # Construction du chemin complet
        if parent_path:
            folder_path = f"{parent_path} / {folder_name}"
        else:
            folder_path = folder_name

        print(f"Analyse du dossier: {folder_path}")

        # Récupération des statistiques
        total_items = 0
        total_size = 0
        total_attachments = 0
        total_attachments_size = 0

        try:
            # Encodage de l'ID du dossier pour l'URL
            encoded_folder_id = urllib.parse.quote(folder_id)
            encoded_user_email = urllib.parse.quote(self.user_email)
            
            # Récupération des messages avec pagination
            url = (
                f"{self.GRAPH_API_ENDPOINT}/users/{encoded_user_email}"
                f"/mailFolders/{encoded_folder_id}/messages"
            )
            
            # Paramètres de la requête
            params = {
                "$select": "id,subject,hasAttachments,size,receivedDateTime",
                "$top": 999,
                "$count": True
            }
            
            headers = {
                **self.session.headers,
                "ConsistencyLevel": "eventual",
                "Prefer": "outlook.allow-unsafe-html"
            }
            
            while url:
                try:
                    response = self.session.get(url, params=params, headers=headers)
                    response.raise_for_status()
                    data = response.json()
                    
                    # Traitement des messages de la page
                    messages = data.get("value", [])
                    for msg in messages:
                        total_items += 1
                        size = msg.get("size", 0)
                        total_size += size
                        
                        if msg.get("hasAttachments", False):
                            total_attachments += 1
                            # Estimation taille PJ (90% taille totale)
                            total_attachments_size += size * 0.9
                    
                    # Passage à la page suivante si elle existe
                    url = data.get("@odata.nextLink")
                    
                    # Mise à jour des paramètres pour la page suivante
                    if url:
                        params = {}  # Les paramètres sont déjà dans nextLink
                        
                except requests.exceptions.RequestException as e:
                    print(f"Erreur réseau pour {folder_path}: {str(e)}")
                    if "429" in str(e):  # Too Many Requests
                        print("Limite d'API atteinte, pause de 60 secondes...")
                        time.sleep(60)
                        continue
                    break

        except Exception as e:
            print(f"Erreur inattendue pour {folder_path}: {str(e)}")
            return None

        # Conversion en Mo
        total_size_mb = total_size / (1024 * 1024)
        total_attachments_size_mb = total_attachments_size / (1024 * 1024)
        
        # Calcul des moyennes
        avg_size_mb = (
            total_size_mb / total_items if total_items > 0 else 0
        )
        avg_attachments_size_mb = (
            total_attachments_size_mb / total_attachments 
            if total_attachments > 0 
            else 0
        )

        # Récupération des actions configurées
        folder_config = self.config.get(folder_id, {})
        inherited_config = None
        inherited_from = None

        # Si pas de config spécifique, chercher dans les parents
        if not folder_config and parent_path:
            parent_folders = parent_path.split(" / ")
            for parent in reversed(parent_folders):
                if parent in self.config:
                    inherited_config = self.config[parent]
                    inherited_from = parent
                    break

        # Utiliser la config héritée si disponible
        config = inherited_config if inherited_config else folder_config

        return FolderStats(
            id=folder_id,
            name=folder_name,
            path=folder_path,
            total_items=total_items,
            total_size_mb=round(total_size_mb, 2),
            total_attachments=total_attachments,
            total_attachments_size_mb=round(total_attachments_size_mb, 2),
            avg_size_mb=round(avg_size_mb, 2),
            avg_attachments_size_mb=round(avg_attachments_size_mb, 2),
            action_clean=config.get("clean", False),
            action_archive=config.get("archive", False),
            action_delete=config.get("delete", False),
            action_min_size_mb=config.get("min_size_mb", 0.0),
            action_max_age_days=config.get("max_age_days", 0),
            inherited_from=inherited_from,
            last_action=config.get("last_action"),
            last_action_date=config.get("last_action_date")
        )

    def analyze_folders(self) -> List[FolderStats]:
        """Analyse récursive des dossiers"""
        stats = []
        folders_to_process = [(self.get_root_folder(), "")]
        
        with tqdm(desc="Analyse des dossiers") as pbar:
            while folders_to_process:
                folder, parent_path = folders_to_process.pop(0)
                
                # Analyse du dossier courant
                folder_stats = self.get_folder_stats(folder, parent_path)
                if folder_stats:
                    stats.append(folder_stats)
                    pbar.update(1)
                
                # Récupération des sous-dossiers
                try:
                    child_folders = self.get_child_folders(folder["id"])
                    for child in child_folders:
                        folders_to_process.append((
                            child,
                            folder_stats.path if folder_stats else ""
                        ))
                except Exception as e:
                    print(
                        f"Erreur lors de la récupération des sous-dossiers "
                        f"de {folder.get('displayName')}: {str(e)}"
                    )
        
        return stats

    def generate_report(self, stats: List[FolderStats]) -> None:
        """Génère le rapport Excel"""
        base_name = "folder_analysis"
        extension = ".xlsx"
        report_file = f"{base_name}{extension}"
        
        # Si le fichier est ouvert, utiliser un nom avec timestamp
        if os.path.exists(report_file):
            try:
                with open(report_file, 'a'):
                    pass
            except PermissionError:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                report_file = f"{base_name}_{timestamp}{extension}"
                print(f"Le fichier {self.REPORT_FILE} est ouvert.")
                print(f"Utilisation du nom alternatif: {report_file}")
        
        print(f"Génération du rapport {report_file}...")
        
        # Création du DataFrame
        df = pd.DataFrame(stats)
        
        # Formatage des colonnes
        df["total_size_mb"] = df["total_size_mb"].round(2)
        df["total_attachments_size_mb"] = df["total_attachments_size_mb"].round(2)
        df["avg_size_mb"] = df["avg_size_mb"].round(2)
        df["avg_attachments_size_mb"] = df["avg_attachments_size_mb"].round(2)
        df["action_min_size_mb"] = df["action_min_size_mb"].round(2)
        
        # Réorganisation et renommage des colonnes
        columns = {
            "name": "Dossier",
            "path": "Chemin",
            "total_items": "Nombre d'éléments",
            "total_size_mb": "Taille totale (Mo)",
            "total_attachments": "Nombre de PJ",
            "total_attachments_size_mb": "Taille PJ (Mo)",
            "avg_size_mb": "Taille moyenne (Mo)",
            "avg_attachments_size_mb": "Taille moyenne PJ (Mo)",
            "action_clean": "Action - Nettoyer",
            "action_archive": "Action - Archiver",
            "action_delete": "Action - Supprimer",
            "action_min_size_mb": "Action - Taille min (Mo)",
            "action_max_age_days": "Action - Age max (jours)",
            "inherited_from": "Hérité de",
            "last_action": "Dernière action",
            "last_action_date": "Date dernière action"
        }
        
        df = df[list(columns.keys())].rename(columns=columns)
        
        # Largeurs minimales pour chaque colonne
        min_widths = {
            "Dossier": 30,
            "Chemin": 40,
            "Nombre d'éléments": 15,
            "Taille totale (Mo)": 15,
            "Nombre de PJ": 12,
            "Taille PJ (Mo)": 15,
            "Taille moyenne (Mo)": 15,
            "Taille moyenne PJ (Mo)": 18,
            "Action - Nettoyer": 15,
            "Action - Archiver": 15,
            "Action - Supprimer": 15,
            "Action - Taille min (Mo)": 18,
            "Action - Age max (jours)": 20,
            "Hérité de": 30,
            "Dernière action": 20,
            "Date dernière action": 20
        }
        
        # Sauvegarde au format Excel avec mise en forme
        max_retries = 3
        retry_delay = 2  # secondes
        
        for attempt in range(max_retries):
            try:
                with pd.ExcelWriter(report_file, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name="Analyse des dossiers")
                    
                    # Mise en forme
                    workbook = writer.book
                    worksheet = writer.sheets["Analyse des dossiers"]
                    
                    # Style pour l'en-tête
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(
                        start_color="366092",
                        end_color="366092",
                        fill_type="solid"
                    )
                    header_alignment = Alignment(
                        horizontal="center",
                        vertical="center",
                        wrap_text=True
                    )
                    
                    # Appliquer le style à l'en-tête
                    for cell in worksheet[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                    
                    # Ajuster la hauteur de la première ligne
                    worksheet.row_dimensions[1].height = 30
                    
                    # Style pour les cellules de données
                    data_alignment = Alignment(
                        horizontal="left",
                        vertical="center"
                    )
                    number_alignment = Alignment(
                        horizontal="right",
                        vertical="center"
                    )
                    
                    # Ajuster la largeur des colonnes et appliquer les styles
                    for idx, column in enumerate(worksheet.columns, 1):
                        column_letter = get_column_letter(idx)
                        header_value = worksheet.cell(row=1, column=idx).value
                        
                        # Calculer la largeur maximale
                        max_length = min_widths.get(header_value, 12)
                        for cell in column[1:]:  # Skip header
                            try:
                                if cell.value:
                                    max_length = max(
                                        max_length,
                                        len(str(cell.value)) + 2
                                    )
                            except:
                                pass
                        
                        # Limiter la largeur maximale
                        max_length = min(max_length, 50)
                        
                        # Appliquer la largeur
                        worksheet.column_dimensions[column_letter].width = max_length
                        
                        # Appliquer l'alignement selon le type de données
                        is_number = any(
                            text in header_value 
                            for text in ["taille", "nombre", "age", "size"]
                        )
                        
                        for cell in column[1:]:  # Skip header
                            cell.alignment = number_alignment if is_number else data_alignment
                    
                    # Ajouter des bordures
                    thin_border = Side(border_style="thin", color="000000")
                    border = Border(
                        left=thin_border,
                        right=thin_border,
                        top=thin_border,
                        bottom=thin_border
                    )
                    
                    for row in worksheet.iter_rows(
                        min_row=1,
                        max_row=worksheet.max_row,
                        min_col=1,
                        max_col=worksheet.max_column
                    ):
                        for cell in row:
                            cell.border = border
                    
                    # Figer la première ligne
                    worksheet.freeze_panes = "A2"
                    
                print(f"Rapport généré avec succès dans {report_file}")
                break
                
            except PermissionError:
                if attempt < max_retries - 1:
                    print(f"Tentative {attempt + 1}/{max_retries} échouée...")
                    print(
                        "Le fichier est peut-être ouvert. "
                        f"Nouvelle tentative dans {retry_delay}s..."
                    )
                    time.sleep(retry_delay)
                    retry_delay *= 2  # Augmente le délai à chaque tentative
                else:
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    new_report_file = f"{base_name}_{timestamp}{extension}"
                    print(
                        "Impossible d'accéder au fichier après "
                        f"{max_retries} tentatives."
                    )
                    print(f"Création d'une nouvelle version: {new_report_file}")
                    
                    # Dernière tentative avec un nouveau nom
                    with pd.ExcelWriter(new_report_file, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False, sheet_name="Analyse des dossiers")
                        print(f"Rapport généré avec succès dans {new_report_file}")
            
            except Exception as e:
                print(f"Erreur lors de la génération du rapport: {str(e)}")
                break

    def set_folder_action(
        self,
        folder_name: str,
        delete_attachments: bool = False,
        max_age_days: int = 365,
        min_size_mb: float = 10.0
    ) -> None:
        """Définit les actions pour un dossier"""
        if folder_name not in self.config:
            self.config[folder_name] = {
                "actions": {},
                "inherited": False,
                "history": []
            }
        
        self.config[folder_name]["actions"] = {
            "delete_attachments": delete_attachments,
            "max_age_days": max_age_days,
            "min_size_mb": min_size_mb
        }
        
        self._save_config()

    def record_action(
        self,
        folder_name: str,
        action: str,
        count: int,
        size_mb: float
    ) -> None:
        """Enregistre une action dans l'historique"""
        if folder_name not in self.config:
            return
        
        history_entry: ActionHistory = {
            "date": datetime.now().strftime("%Y-%m-%d"),
            "action": action,
            "count": count,
            "size_mb": size_mb
        }
        
        self.config[folder_name]["history"].append(history_entry)
        self._save_config()

    def get_root_folder(self) -> Dict:
        """Récupère le dossier racine"""
        try:
            # Utiliser le dossier "inbox" comme point de départ
            url = (
                f"{self.GRAPH_API_ENDPOINT}/users/{self.user_email}"
                f"/mailFolders/inbox"
            )
            response = self.session.get(url)
            response.raise_for_status()
            inbox = response.json()
            
            # Remonter au dossier parent
            if "parentFolderId" in inbox:
                parent_id = inbox["parentFolderId"]
                url = (
                    f"{self.GRAPH_API_ENDPOINT}/users/{self.user_email}"
                    f"/mailFolders/{parent_id}"
                )
                response = self.session.get(url)
                response.raise_for_status()
                return response.json()
            
            return inbox
            
        except Exception as e:
            print(f"Erreur lors de la récupération du dossier racine: {str(e)}")
            raise

    def get_child_folders(self, folder_id: str) -> List[Dict]:
        """Récupère les sous-dossiers d'un dossier"""
        try:
            # Encoder l'ID du dossier
            encoded_folder_id = urllib.parse.quote(folder_id)
            encoded_user_email = urllib.parse.quote(self.user_email)
            
            url = (
                f"{self.GRAPH_API_ENDPOINT}/users/{encoded_user_email}"
                f"/mailFolders/{encoded_folder_id}/childFolders"
            )
            
            # Ajouter les paramètres de sélection
            params = {
                "$select": "id,displayName,parentFolderId,childFolderCount,unreadItemCount,totalItemCount",
                "$top": 999
            }
            
            response = self.session.get(url, params=params)
            response.raise_for_status()
            return response.json().get("value", [])
            
        except Exception as e:
            print(
                f"Erreur lors de la récupération des sous-dossiers: {str(e)}"
            )
            return []


def main():
    """Point d'entrée principal"""
    analyzer = OutlookFolderAnalyzer()
    
    # Analyse des dossiers
    stats = analyzer.analyze_folders()
    
    # Génération du rapport
    analyzer.generate_report(stats)


if __name__ == "__main__":
    main() 