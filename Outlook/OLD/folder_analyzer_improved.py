#!/usr/bin/env python3
"""
Script d'analyse des dossiers Outlook avec l'API Microsoft Graph.
Combine les fonctionnalités de diagnostic_emails.py et folder_analyzer.py.
"""

import os
import sys
import logging
from typing import Dict, List, NamedTuple, Optional

import requests
import pandas as pd
from dotenv import load_dotenv
from azure.identity import ClientSecretCredential
from openpyxl.styles import (
    Font, PatternFill, Alignment, Side, Border
)
from openpyxl.utils import get_column_letter
from outlook_cleaner_improved import get_messages_improved


class FolderStats(NamedTuple):
    """Structure pour les statistiques d'un dossier"""
    id: str
    name: str
    path: str
    total_items: int
    total_size_mb: float
    total_attachments: int
    total_attachments_size_mb: float
    avg_size_mb: float
    avg_attachments_size_mb: float
    last_message_date: Optional[str]


def get_access_token():
    """Obtient un token d'accès pour l'API Microsoft Graph."""
    try:
        client_id = os.getenv('OUTLOOK_CLIENT_ID')
        client_secret = os.getenv('OUTLOOK_CLIENT_SECRET')
        tenant_id = os.getenv('TENANT_ID')
        
        # Afficher les variables (masquées pour la sécurité)
        print(f"Client ID présent: {'Oui' if client_id else 'Non'}")
        print(f"Client Secret présent: {'Oui' if client_secret else 'Non'}")
        print(f"Tenant ID présent: {'Oui' if tenant_id else 'Non'}")

        if not all([client_id, client_secret, tenant_id]):
            raise ValueError(
                "Les variables d'environnement OUTLOOK_CLIENT_ID, "
                "OUTLOOK_CLIENT_SECRET et TENANT_ID doivent être définies."
            )

        print("Tentative d'authentification avec Azure...")
        credentials = ClientSecretCredential(
            tenant_id=tenant_id,
            client_id=client_id,
            client_secret=client_secret
        )

        print("Obtention du token...")
        token = credentials.get_token("https://graph.microsoft.com/.default")
        print("Token obtenu avec succès!")
        return token.token
    except Exception as e:
        logging.error(
            f"Erreur détaillée lors de l'obtention du token: {str(e)}"
        )
        raise


class OutlookAnalyzer:
    """Classe pour analyser les dossiers Outlook"""

    GRAPH_API_ENDPOINT = "https://graph.microsoft.com/v1.0"
    REPORT_FILE = "outlook_analysis.xlsx"

    def __init__(self):
        """Initialisation de l'analyseur"""
        # Configuration du logging
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s'
        )

        # Chargement des variables d'environnement
        load_dotenv()
        self.user_email = os.getenv("OUTLOOK_USER_EMAIL")
        if not self.user_email:
            raise ValueError("OUTLOOK_USER_EMAIL doit être défini")

        # Création de la session avec le token
        self.session = requests.Session()
        token = get_access_token()
        self.session.headers.update({
            "Authorization": f"Bearer {token}",
            "Accept": "application/json"
        })

    def get_folder_stats(self, folder: Dict, parent_path: str = "") -> FolderStats:
        """Récupère les statistiques d'un dossier"""
        folder_id = folder["id"]
        folder_name = folder["displayName"]
        
        # Construction du chemin complet
        if parent_path:
            folder_path = f"{parent_path} / {folder_name}"
        else:
            folder_path = folder_name

        print(f"Analyse du dossier: {folder_path}")

        try:
            # Utilisation de la fonction qui marche
            messages = get_messages_improved(
                self.user_email,
                folder_id=folder_id,
                limit=50  # On garde la même limite pour l'instant
            )
            
            if messages is None:
                return None
                
            total_items = len(messages)
            total_size = sum(msg.get("size", 0) for msg in messages)
            total_attachments = sum(1 for msg in messages if msg.get("hasAttachments", False))
            
            # Calcul des moyennes et tailles en Mo
            total_size_mb = total_size / (1024 * 1024)
            avg_size_mb = total_size_mb / total_items if total_items > 0 else 0

            # Constante pour le chemin à remplacer
            old_path = "Partie supérieure de la banque d'informations / "

            return FolderStats(
                id=folder_id,
                name=folder_name,
                path=folder_path.replace(old_path, ""),
                total_items=total_items,
                total_size_mb=total_size_mb,
                total_attachments=total_attachments,
                total_attachments_size_mb=0,  # Temporairement désactivé
                avg_size_mb=avg_size_mb,
                avg_attachments_size_mb=0,  # Temporairement désactivé
                last_message_date=None  # Temporairement désactivé
            )

        except Exception as e:
            err_msg = f"Erreur inattendue pour {folder_path}: {str(e)}"
            print(err_msg)
            return None

    def get_root_folder(self) -> Dict:
        """Récupère le dossier racine"""
        try:
            url = (
                f"{self.GRAPH_API_ENDPOINT}"
                f"/users/{self.user_email}"
                "/mailFolders"
            )
            
            params = {
                "$select": "id,displayName,parentFolderId,childFolderCount",
                "$top": 999,
                "$filter": "childFolderCount gt 0"
            }
            
            response = self.session.get(url, params=params)
            response.raise_for_status()
            folders = response.json().get("value", [])
            
            # Trouver le dossier racine (celui qui a le plus de sous-dossiers)
            root_folder = max(
                folders,
                key=lambda f: f.get("childFolderCount", 0),
                default=None
            )
            
            if not root_folder:
                raise Exception("Dossier racine non trouvé")
                
            return root_folder
            
        except Exception as e:
            print(
                f"Erreur lors de la récupération du dossier racine: {str(e)}"
            )
            raise

    def get_child_folders(self, folder: Dict) -> List[Dict]:
        """Récupère les sous-dossiers d'un dossier"""
        try:
            url = (
                f"{self.GRAPH_API_ENDPOINT}"
                f"/users/{self.user_email}"
                f"/mailFolders/{folder['id']}/childFolders"
            )
            
            params = {
                "$select": "id,displayName,parentFolderId",
                "$top": 999
            }
            
            response = self.session.get(url, params=params)
            response.raise_for_status()
            return response.json().get("value", [])
            
        except Exception as e:
            print(
                f"Erreur lors de la récupération des sous-dossiers: {str(e)}"
            )
            return []

    def analyze_folders(self) -> List[FolderStats]:
        """Analyse récursivement tous les dossiers"""
        all_stats = []
        folders_to_process = [(self.get_root_folder(), "")]
        
        while folders_to_process:
            folder, parent_path = folders_to_process.pop(0)
            
            # Analyse du dossier courant
            stats = self.get_folder_stats(folder, parent_path)
            if stats:
                all_stats.append(stats)
            
            # Récupération des sous-dossiers
            child_folders = self.get_child_folders(folder)
            folders_to_process.extend(
                (f, stats.path if stats else parent_path)
                for f in child_folders
            )
        
        return all_stats

    def generate_report(self, stats: List[FolderStats]) -> None:
        """Génère un rapport Excel avec les statistiques"""
        # Création du DataFrame
        df = pd.DataFrame([
            {
                "Dossier": s.path.replace(
                    "Partie supérieure de la banque d'informations / ", ""
                ),
                "Nombre d'éléments": s.total_items,
                "Taille totale (Mo)": round(s.total_size_mb, 2),
                "Nombre de pièces jointes": s.total_attachments,
                "Taille PJ (Mo)": round(s.total_attachments_size_mb, 2),
                "Taille moyenne (Mo)": round(s.avg_size_mb, 2),
                "Taille moyenne PJ (Mo)": round(s.avg_attachments_size_mb, 2),
                "Dernier message": s.last_message_date or "N/A",
                "Action": "",
                "Commentaire": "",
                "Statut": ""
            }
            for s in stats
        ])

        # Tri par taille totale décroissante
        df = df.sort_values("Taille totale (Mo)", ascending=False)

        # Sauvegarde en Excel
        writer = pd.ExcelWriter(self.REPORT_FILE, engine="openpyxl")
        df.to_excel(writer, index=False, sheet_name="Analyse")
        
        # Récupération de la feuille active
        wb = writer.book
        ws = wb.active
        
        # Définition des styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Bordures
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # Application des styles aux en-têtes
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Formatage des colonnes
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            ws.column_dimensions[letter].width = 15
            
            # Ajustements spécifiques
            if col == 1:  # Colonne Dossier
                ws.column_dimensions[letter].width = 40
            elif col in [3, 5]:  # Colonnes de taille
                for cell in ws[letter][1:]:
                    if cell.value and isinstance(cell.value, (int, float)):
                        cell.number_format = "#,##0.00"
            
            # Bordures et alignement pour toutes les cellules
            for cell in ws[letter][1:]:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Alignement à gauche pour la colonne Dossier
                if col == 1:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Ajustement automatique de la hauteur des lignes
        for row in ws.rows:
            max_height = 0
            for cell in row:
                if cell.value:
                    text_height = len(str(cell.value).split('\n')) * 15
                    max_height = max(max_height, text_height)
            ws.row_dimensions[cell.row].height = max_height
        
        # Gel de la première ligne
        ws.freeze_panes = "A2"
        
        # Sauvegarde du fichier
        writer.close()
        print(f"\nRapport généré dans {self.REPORT_FILE}")


def main():
    """Point d'entrée principal"""
    try:
        analyzer = OutlookAnalyzer()
        print("\nAnalyse des dossiers en cours...")
        stats = analyzer.analyze_folders()
        print("\nGénération du rapport...")
        analyzer.generate_report(stats)
        print("\nAnalyse terminée avec succès!")
        
    except Exception as e:
        print(f"\nErreur lors de l'analyse: {str(e)}")
        sys.exit(1)


if __name__ == "__main__":
    main()